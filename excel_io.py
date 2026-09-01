# -*- coding: utf-8 -*-
"""
IrwaneTraceForest (ITF) - Import / Export Excel
Générique : fonctionne pour n'importe quel module (inventaire, production,
stock, audit) à partir d'une liste d'en-têtes et de lignes de données.
"""

import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ENTETE_FOND = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
ENTETE_POLICE = Font(color="FFFFFF", bold=True, size=10)


def exporter_xlsx(titre: str, entetes: list, lignes: list) -> bytes:
    """entetes : liste de libellés de colonnes.
    lignes : liste de tuples/listes, dans le même ordre que entetes."""
    wb = Workbook()
    ws = wb.active
    ws.title = titre[:31] if titre else "Export"

    for col_idx, libelle in enumerate(entetes, start=1):
        cell = ws.cell(row=1, column=col_idx, value=libelle)
        cell.font = ENTETE_POLICE
        cell.fill = ENTETE_FOND
        cell.alignment = Alignment(horizontal="center")

    for row_idx, ligne in enumerate(lignes, start=2):
        for col_idx, valeur in enumerate(ligne, start=1):
            ws.cell(row=row_idx, column=col_idx, value=valeur)

    for col_idx, libelle in enumerate(entetes, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(str(libelle)) + 4)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importer_xlsx(fichier_bytes: bytes, colonnes_attendues: list) -> list:
    """Lit un fichier .xlsx et retourne une liste de dictionnaires
    {colonne: valeur}, en s'appuyant sur l'ordre des colonnes_attendues
    (première ligne = en-têtes, ignorée pour le mapping mais vérifiée).
    Lève ValueError si le fichier est vide ou illisible."""
    try:
        wb = load_workbook(io.BytesIO(fichier_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Fichier Excel illisible : {exc}")

    ws = wb.active
    lignes = list(ws.iter_rows(values_only=True))
    if len(lignes) < 2:
        return []

    resultats = []
    for ligne in lignes[1:]:
        if all(v is None for v in ligne):
            continue
        enregistrement = {}
        for i, col in enumerate(colonnes_attendues):
            enregistrement[col] = ligne[i] if i < len(ligne) else None
        resultats.append(enregistrement)
    return resultats
